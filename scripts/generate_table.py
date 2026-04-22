#!/usr/bin/env python3
"""Генерация реестра документов из .vassal/index.yaml."""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml


OUTPUT_XLSX = "Таблица документов.xlsx"
OUTPUT_CSV = "Таблица документов.csv"
HEADERS = [
    "№",
    "Дата",
    "Тип",
    "Отправитель -> Адресат",
    "Название",
    "Краткое содержание",
    "Печать/Подпись",
    "Полнота",
    "Качество",
    "Файл",
]
GROUP_ORDER = [
    "Договоры",
    "Переписка",
    "Процессуальные",
    "Платёжные документы",
    "Прочие",
]
GROUP_MAP = {
    "договор": "Договоры",
    "дополнительное-соглашение": "Договоры",
    "приложение-к-договору": "Договоры",
    "спецификация": "Договоры",
    "акт": "Договоры",
    "накладная": "Договоры",
    "счёт": "Договоры",
    "счет": "Договоры",
    "счёт-фактура": "Договоры",
    "счет-фактура": "Договоры",
    "переписка": "Переписка",
    "претензия": "Переписка",
    "ответ-на-претензию": "Переписка",
    "уведомление": "Переписка",
    "письмо": "Переписка",
    "требование": "Переписка",
    "процессуальные": "Процессуальные",
    "иск": "Процессуальные",
    "исковое-заявление": "Процессуальные",
    "отзыв": "Процессуальные",
    "отзыв-на-иск": "Процессуальные",
    "ходатайство": "Процессуальные",
    "заявление": "Процессуальные",
    "возражение": "Процессуальные",
    "пояснения": "Процессуальные",
    "дополнение-к-позиции": "Процессуальные",
    "апелляционная-жалоба": "Процессуальные",
    "кассационная-жалоба": "Процессуальные",
    "определение": "Процессуальные",
    "решение": "Процессуальные",
    "постановление": "Процессуальные",
    "протокол-заседания": "Процессуальные",
    "платёжные": "Платёжные документы",
    "платежные": "Платёжные документы",
    "платёжное-поручение": "Платёжные документы",
    "платежное-поручение": "Платёжные документы",
    "выписка": "Платёжные документы",
    "справка-из-банка": "Платёжные документы",
}


@dataclass
class Row:
    values: list[str]
    row_type: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Сгенерировать Таблица документов.xlsx из .vassal/index.yaml"
    )
    parser.add_argument(
        "--case-root",
        required=True,
        help="Абсолютный или относительный путь к папке дела",
    )
    return parser.parse_args()


def error(message: str) -> int:
    print(f"ERROR: {message}")
    return 1


def format_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return text


def bool_label(value: Any) -> str:
    if value is True:
        return "да"
    if value is False:
        return "нет"
    if value in (None, ""):
        return ""
    return str(value)


def compose_seal_signature(doc: dict[str, Any]) -> str:
    parts: list[str] = []
    seal = bool_label(doc.get("seal"))
    signature = bool_label(doc.get("signature"))
    if seal:
        parts.append(f"Печать: {seal}")
    if signature:
        parts.append(f"Подпись: {signature}")
    return "; ".join(parts)


def normalize_parties(parties: Any) -> str:
    if parties in (None, ""):
        return ""
    if isinstance(parties, str):
        return parties
    if isinstance(parties, dict):
        sender = parties.get("from") or parties.get("sender") or ""
        recipient = parties.get("to") or parties.get("recipient") or ""
        if sender or recipient:
            return f"{sender} -> {recipient}".strip()
        return ", ".join(str(value) for value in parties.values() if value)
    if isinstance(parties, list):
        rendered: list[str] = []
        for item in parties:
            if isinstance(item, dict):
                sender = item.get("from") or item.get("sender") or ""
                recipient = item.get("to") or item.get("recipient") or ""
                if sender or recipient:
                    rendered.append(f"{sender} -> {recipient}".strip())
                else:
                    rendered.append(
                        ", ".join(str(value) for value in item.values() if value)
                    )
            elif item not in (None, ""):
                rendered.append(str(item))
        return "; ".join(part for part in rendered if part)
    return str(parties)


def doc_type_value(doc: dict[str, Any]) -> str:
    return str(doc.get("doc_type") or doc.get("type") or "").strip()


def group_for_doc(doc: dict[str, Any]) -> str:
    doc_type = doc_type_value(doc).lower()
    return GROUP_MAP.get(doc_type, "Прочие")


def build_row(number: int, doc: dict[str, Any], title_override: str | None = None) -> Row:
    title = title_override if title_override is not None else str(doc.get("title") or "")
    values = [
        str(number),
        format_date(doc.get("date")),
        doc_type_value(doc),
        normalize_parties(doc.get("parties")),
        title,
        str(doc.get("summary") or ""),
        compose_seal_signature(doc),
        str(doc.get("completeness") or ""),
        str(doc.get("quality") or ""),
        str(doc.get("file") or ""),
    ]
    return Row(values=values, row_type="document")


def resolve_member_ids(doc: dict[str, Any]) -> list[str]:
    members = doc.get("bundle_members") or []
    result: list[str] = []
    if isinstance(members, list):
        for item in members:
            if isinstance(item, str):
                result.append(item)
            elif isinstance(item, dict) and item.get("id"):
                result.append(str(item["id"]))
    return result


def flatten_documents(documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    doc_by_id = {str(doc.get("id")): doc for doc in documents if doc.get("id")}
    index_by_id = {
        str(doc.get("id")): idx for idx, doc in enumerate(documents) if doc.get("id")
    }
    bundle_members: dict[str, list[str]] = {}
    bundle_anchors: dict[str, str] = {}

    for doc in documents:
        doc_id = str(doc.get("id") or "")
        if not doc_id:
            continue
        for member_id in resolve_member_ids(doc):
            bundle_members.setdefault(doc_id, []).append(member_id)
        bundle_id = doc.get("bundle_id")
        if not bundle_id:
            continue
        bundle_id = str(bundle_id)
        if doc.get("anchor") is True or doc.get("role_in_bundle") == "main":
            bundle_anchors[bundle_id] = doc_id
        elif doc.get("member") is True:
            bundle_members.setdefault(bundle_id, []).append(doc_id)

    processed: set[str] = set()
    flattened: list[dict[str, Any]] = []

    def member_sort_key(member_id: str) -> int:
        return index_by_id.get(member_id, 10**9)

    for doc in documents:
        doc_id = str(doc.get("id") or "")
        if not doc_id or doc_id in processed:
            continue

        if doc.get("member") is True and doc.get("bundle_id"):
            bundle_id = str(doc["bundle_id"])
            anchor_id = bundle_anchors.get(bundle_id)
            if anchor_id and anchor_id != doc_id:
                continue

        flattened.append(doc)
        processed.add(doc_id)

        explicit_members = bundle_members.get(doc_id, []).copy()
        bundle_id = doc.get("bundle_id")
        if bundle_id and (doc.get("anchor") is True or doc.get("role_in_bundle") == "main"):
            explicit_members.extend(bundle_members.get(str(bundle_id), []))

        seen_member_ids: set[str] = set()
        for member_id in sorted(explicit_members, key=member_sort_key):
            if member_id in seen_member_ids:
                continue
            seen_member_ids.add(member_id)
            member_doc = doc_by_id.get(member_id)
            if not member_doc:
                continue
            flattened.append(
                {
                    **member_doc,
                    "_bundle_child": True,
                    "_display_title": f"  \u2192 {member_doc.get('title') or ''}",
                }
            )
            processed.add(member_id)

    for doc in documents:
        doc_id = str(doc.get("id") or "")
        if doc_id and doc_id not in processed:
            flattened.append(doc)
            processed.add(doc_id)

    return flattened


def build_rows(documents: list[dict[str, Any]]) -> list[Row]:
    grouped: dict[str, list[dict[str, Any]]] = {group: [] for group in GROUP_ORDER}
    for doc in flatten_documents(documents):
        grouped[group_for_doc(doc)].append(doc)

    rows: list[Row] = []
    number = 1
    for group in GROUP_ORDER:
        docs = grouped[group]
        if not docs:
            continue
        rows.append(Row(values=[group] + [""] * (len(HEADERS) - 1), row_type="group"))
        for doc in docs:
            title_override = doc.get("_display_title")
            rows.append(build_row(number, doc, title_override=title_override))
            number += 1
    return rows


def write_csv(output_path: Path, rows: list[Row]) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow(row.values)


def write_xlsx(output_path: Path, rows: list[Row]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр документов"
    sheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    group_font = Font(bold=True, italic=True)

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = header_font

    current_row = 2
    for row in rows:
        sheet.append(row.values)
        if row.row_type == "group":
            for cell in sheet[current_row]:
                cell.font = group_font
        current_row += 1

    for index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        letter = get_column_letter(index)
        for cell in sheet[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value.replace("\n", " ")))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 80)

    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    case_root = Path(args.case_root).expanduser().resolve()
    index_path = case_root / ".vassal" / "index.yaml"

    if not case_root.exists():
        return error(f"папка дела не найдена: {case_root}")
    if not index_path.exists():
        return error(f"не найден index.yaml: {index_path}")

    try:
        data = yaml.safe_load(index_path.read_text(encoding="utf-8")) or {}
    except Exception as exc:  # pragma: no cover - defensive path
        return error(f"не удалось прочитать index.yaml: {exc}")

    documents = data.get("documents")
    if not isinstance(documents, list):
        return error("поле documents отсутствует или не является списком")

    rows = build_rows(documents)
    document_count = len(documents)
    xlsx_path = case_root / OUTPUT_XLSX

    try:
        write_xlsx(xlsx_path, rows)
        print(f"OK: {document_count} документов записано в {OUTPUT_XLSX}")
        return 0
    except ModuleNotFoundError as exc:
        if exc.name != "openpyxl":
            return error(str(exc))
        csv_path = case_root / OUTPUT_CSV
        try:
            write_csv(csv_path, rows)
        except Exception as csv_exc:  # pragma: no cover - defensive path
            return error(f"не удалось записать CSV fallback: {csv_exc}")
        print(
            f"WARNING: openpyxl не установлен, вместо {OUTPUT_XLSX} записан {OUTPUT_CSV}"
        )
        print(f"OK: {document_count} документов записано в {OUTPUT_CSV}")
        return 0
    except Exception as exc:  # pragma: no cover - defensive path
        return error(str(exc))


if __name__ == "__main__":
    sys.exit(main())
